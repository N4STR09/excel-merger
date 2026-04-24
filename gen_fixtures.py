"""
Genera los fixtures Excel usados por los tests:
  - src/test/resources/fixtures/cierre.xlsx    (peticiones del ERP)
  - src/test/resources/fixtures/extraccion.xlsx (export de Jira)

Se deja este script para poder regenerar los fixtures si cambian las
cabeceras esperadas por los perfiles. Los .xlsx resultantes se
versionan en src/test/resources/fixtures/ junto con los tests.

v2.0.0: los nombres de perfil se intercambiaron respecto a v1.x. Los
ficheros xlsx se renombran tambien para que el nombre de fichero
coincida con el nombre del perfil que los detecta por contenido:
  - cierre.xlsx    contiene cabeceras del perfil Cierre    (peticiones ERP)
  - extraccion.xlsx contiene cabeceras del perfil Extraccion (export Jira)

Uso (desde la raiz del proyecto):
    pip install openpyxl --break-system-packages
    python3 gen_fixtures.py

v1.6.2: se incorporan 3 filas de regresion para verificar que el SUMIFS
tolera mismatch de tipo numero/texto entre los dos perfiles.

v1.8.0: se anade una fila extra con un responsable escrito en MAYUSCULAS
(TRESP1@x) para validar que la nueva tabla "Totales Peticiones por
Responsables Matriculas" normaliza correctamente (tresp1@x y TRESP1@x
cuentan como un unico responsable TRESP1@X, y el SUMIFS case-insensitive
suma ambas variantes).

v1.8.1: se anade una fila con Usuario_Resp_Tecnico="MG002   " (padding
de espacios) para validar que el trim en la capa de copia hace casar
correctamente el SUMIFS de la segunda tabla de Resumen.
"""

from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell  # noqa: F401  (util potencial)

OUT_DIR = "src/test/resources/fixtures"

# =========================================================================
# cierre.xlsx — peticiones del ERP
# =========================================================================
# Perfil Cierre (v2.0.0; antes Extraccion): cabeceras en fila 1. Debe
# tener al menos 4 cabeceras de las del profile para que se detecte.
# 14 filas "historicas" (todas texto) + 3 filas de regresion v1.6.2
# (Peticion/Recurso como NUMERIC) + 1 regresion v1.8.0 + 1 regresion
# v1.8.1 + 1 skip = 20 filas utiles + 1 header.
CIERRE_PROFILE_HEADERS = [
    "Peticion", "Titulo", "Aplicaci_Activi", "Estado", "Usuario_Resp_Tecnico",
    "Horas_AutoriTotPeticion", "Horas_RealizadoTot", "Estado_Distribucion",
    "Planificacion", "Objeto_EstudioPeticion", "Codigo_Facturacion",
    "Recurso", "Funcion", "UltimaPrevision_Horas_Mes", "Realizadas_Horas_Mes",
    "Total_Horas_Autorizadas_Recurso", "Total_Horas_Realizadas_Recurso",
]

def build_cierre_profile_fixture():
    """Genera cierre.xlsx (perfil Cierre v2.0.0 = peticiones ERP)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cierre"
    ws.append(CIERRE_PROFILE_HEADERS)

    # 14 filas validas (texto en Peticion y Recurso)
    rows = [
        ["P-001", "Migracion BBDD",          "DF", "Abierta",  "tresp1@x", 40, 12, "OK", "P1", "Obj A", "CTR-100", "M-1001", "Dev",    10, 5,  40, 20],
        ["P-002", "Refactor UI",             "HE", "Abierta",  "tresp1@x", 60, 30, "OK", "P1", "Obj B", "CTR-100", "M-1002", "Dev",    20, 15, 60, 35],
        ["P-003", "API v2",                  "EW", "Cerrada",  "tresp2@x", 25, 25, "OK", "P2", "Obj C", "CTR-200", "M-1003", "Dev",     0,  0, 25, 25],
        ["P-004", "Dashboards",              "J6", "Abierta",  "tresp2@x", 80, 10, "OK", "P2", "Obj D", "CTR-200", "M-1004", "Dev",    30, 10, 80, 20],
        ["P-005", "Integracion SSO",         "JX", "Abierta",  "tresp1@x", 45,  0, "OK", "P1", "Obj E", "CTR-100", "M-1001", "Dev",    15,  0, 45,  5],
        ["P-006", "Alertas",                 "HW", "Abierta",  "tresp3@x", 20,  8, "OK", "P3", "Obj F", "CTR-300", "M-1005", "Dev",     5,  2, 20, 10],
        ["P-007", "Migracion cloud",         "DF", "Abierta",  "tresp1@x", 100, 40,"OK", "P1", "Obj G", "CTR-100", "M-1001", "Dev",    50, 15,100, 60],
        ["P-008", "Seguridad",               "HE", "Cerrada",  "tresp2@x", 30, 30, "OK", "P2", "Obj H", "CTR-200", "M-1002", "Dev",     0,  0, 30, 30],
        ["P-009", "Perf mejora",             "EW", "Abierta",  "tresp2@x", 35, 20, "OK", "P2", "Obj I", "CTR-200", "M-1003", "Dev",    10,  5, 35, 25],
        ["P-010", "Tooling",                 "PE", "Abierta",  "tresp3@x", 12,  4, "OK", "P3", "Obj J", "CTR-300", "M-1006", "Dev",     4,  1, 12,  4],
        ["P-011", "ZZ-App nueva (sin mapeo)","ZZ", "Abierta",  "tresp3@x", 18,  6, "OK", "P3", "Obj K", "CTR-300", "M-1007", "Dev",     6,  2, 18,  7],
        ["P-012", "Refactor tests",          "WH", "Abierta",  "tresp1@x", 22, 10, "OK", "P1", "Obj L", "CTR-100", "M-1002", "Dev",     8,  3, 22, 11],
        ["P-013", "Import CSV",              "PF", "Abierta",  "tresp2@x", 15,  5, "OK", "P2", "Obj M", "CTR-200", "M-1008", "Dev",     5,  1, 15,  6],
        ["P-014", "Export Parquet",          "KE", "Abierta",  "tresp3@x", 28, 12, "OK", "P3", "Obj N", "CTR-300", "M-1006", "Dev",     7,  3, 28, 13],
    ]
    for r in rows:
        ws.append(r)

    # Filas de regresion v1.6.2 — mismatch de tipos numerico/textual.
    # Peticion y Recurso vienen como NUMERIC aqui, tal y como llegan en
    # el export real del usuario. openpyxl los serializa como numeric.
    # En el fichero del perfil Extraccion (v2.0.0; antes perfil Cierre),
    # las imputaciones para estas peticiones/matriculas vienen como
    # STRING. Sin el fix v1.6.2, el SUMIFS daria 0 para las tres.
    regression_rows = [
        # Peticion=55751 (num), Recurso=99642 (num). En el fichero del
        # perfil Extraccion habra una imputacion con Component Name="55751"
        # (str), Matricula="99642" (str), Funcion="Dev", Hours=7.
        # Esperado tras fix: Jira=7.
        [55751, "Regresion num-num", "DF", "Abierta",  "tresp1@x", 50, 7, "OK", "P1", "Obj", "CTR-100", 99642, "Dev", 25, 7, 50, 7],
        # Peticion=101770 (num), Recurso=90014 (num). Dos imputaciones en
        # el fichero del perfil Extraccion (3 + 2 = 5). Esperado tras
        # fix: Jira=5.
        [101770, "Regresion num dobles", "EW", "Abierta",  "tresp2@x", 40, 5, "OK", "P2", "Obj", "CTR-200", 90014, "Dev", 20, 5, 40, 5],
        # Peticion=138074 (num), Recurso=99641 (num). Una imputacion de 9h
        # + una de 4h con Funcion=Sup (filtrada por el tercer criterio).
        # Esperado tras fix: Jira=9 (solo la Dev).
        [138074, "Regresion num filtrada Sup", "HE", "Abierta",  "tresp1@x", 30, 9, "OK", "P1", "Obj", "CTR-100", 99641, "Dev", 15, 9, 30, 9],
    ]
    for r in regression_rows:
        ws.append(r)

    # Filas de regresion v1.8.0 — normalizacion de responsables en la
    # segunda tabla de Resumen. Responsable escrito en MAYUSCULAS; debe
    # colapsar con "tresp1@x" en la tabla "Totales Peticiones por
    # Responsables Matriculas" (una sola columna "TRESP1@X" con el
    # SUMIFS case-insensitive sumando ambas variantes).
    #
    # Peticion=P-015 (str), Recurso=M-1009 (str), Responsable en MAYUS,
    # PDCL esperado = Jira*1.2 = 0*1.2 = 0 (no tiene imputacion en el
    # fichero del perfil Extraccion, asi que solo verifica el agrupamiento).
    # Para verificar que suma efectivamente, el test usa tresp1@x que si
    # tiene horas.
    responsible_case_row = [
        "P-015", "Regresion responsable MAYUS", "DF", "Abierta",
        "TRESP1@x",   # mismo responsable que tresp1@x pero en MAYUS
        20, 0, "OK", "P1", "Obj V", "CTR-100", "M-1009", "Dev",
        5, 0, 20, 5,
    ]
    ws.append(responsible_case_row)

    # Fila de regresion v1.8.1 — padding de espacios en el responsable.
    # El export real del ERP alinea codigos a 8 caracteres con espacios
    # a la derecha ("MG002   "). Sin el trim en la capa de copia, el
    # SUMIFS de la segunda tabla de Resumen comparaba "MG002" (criterio
    # normalizado en la cabecera) contra "MG002   " (celda con padding)
    # y no casaba — todos los valores salian 0 en Excel real. Con trim
    # activo debe sumar correctamente.
    #
    # Peticion=P-016, Recurso=M-1010, Responsable="MG002   " (padding).
    # Imputacion correspondiente en el fichero del perfil Extraccion:
    # PROJ-40 con Component Name "P-016", Matricula "M-1010",
    # Funcion "Dev", Hours=5. Esperado tras el trim: Jira=5, PDCL=6.0.
    # En la tabla por responsable, la celda
    # ("M-1010", "MG002") debe ser 6.0 (el PDCL, que es la metrica que
    # usa summary.byResponsible).
    responsible_padded_row = [
        "P-016", "Regresion responsable con padding", "DF", "Abierta",
        "MG002   ",   # 3 espacios al final — caso real del export ERP
        20, 5, "OK", "P1", "Obj W", "CTR-100", "M-1010", "Dev",
        5, 5, 20, 5,
    ]
    ws.append(responsible_padded_row)

    # Ultima fila: Peticion vacia → debe ser saltada por MesSheetBuilder (ancla vacia)
    skip_row = [""] + ["-"] * (len(CIERRE_PROFILE_HEADERS) - 1)
    ws.append(skip_row)

    wb.save(f"{OUT_DIR}/cierre.xlsx")
    print("Generated cierre.xlsx (perfil Cierre v2.0.0 = peticiones ERP): "
          "1 header + 14 data + 3 regression v1.6.2 + 1 regression v1.8.0 + "
          "1 regression v1.8.1 + 1 skip = 21 filas")


# =========================================================================
# extraccion.xlsx — export de Jira
# =========================================================================
# Perfil Extraccion (v2.0.0; antes Cierre): cabeceras en fila 2 (la 1 es
# titulo). 16 imputaciones historicas + 5 imputaciones de regresion v1.6.2/
# v1.8.1 para cruzar con las filas numericas y las de padding anadidas al
# perfil Cierre.
EXTRACCION_PROFILE_HEADERS = [
    "Project Key", "Issue Key", "Aplicación BFA", "Labels",
    "Parent Issue Summary", "Summary", "Description", "Issue Type",
    "Time entry: User", "Time entry: Date", "Component Name",
    "Time entry: Description", "Matricula", "Funcion", "Account", "Hours",
]

def build_extraccion_profile_fixture():
    """Genera extraccion.xlsx (perfil Extraccion v2.0.0 = export Jira)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extraccion"

    # Fila 1: metadata / titulo
    ws.append(["EXPORT JIRA - Closure Report"] + [None] * (len(EXTRACCION_PROFILE_HEADERS) - 1))
    # Fila 2: cabeceras
    ws.append(EXTRACCION_PROFILE_HEADERS)

    # 16 filas historicas (texto en todas las columnas). Totales conocidos
    # (Funcion=Dev):
    #   P-001 + M-1001 + Dev -> PROJ-1 (3) + PROJ-2 (2) = 5 horas   (PROJ-3 con Sup queda fuera)
    #   P-002 + M-1002 + Dev -> PROJ-4 (5) = 5 horas
    #   P-003 + M-1003 + Dev -> PROJ-5 (10) + PROJ-6 (5) = 15 horas
    rows = [
        ["PROJ", "PROJ-1",  "DF", "back", "epic-1",  "sum1",  "desc",  "Task", "u1", "2026-01-02", "P-001", "t1", "M-1001", "Dev", "acc1",  3],
        ["PROJ", "PROJ-2",  "DF", "back", "epic-1",  "sum2",  "desc",  "Task", "u1", "2026-01-03", "P-001", "t2", "M-1001", "Dev", "acc1",  2],
        ["PROJ", "PROJ-3",  "DF", "back", "epic-1",  "sum3",  "desc",  "Task", "u1", "2026-01-04", "P-001", "t3", "M-1001", "Sup", "acc1",  4],
        ["PROJ", "PROJ-4",  "HE", "ui",   "epic-2",  "sum4",  "desc",  "Task", "u2", "2026-01-05", "P-002", "t4", "M-1002", "Dev", "acc1",  5],
        ["PROJ", "PROJ-5",  "EW", "api",  "epic-3",  "sum5",  "desc",  "Task", "u3", "2026-01-06", "P-003", "t5", "M-1003", "Dev", "acc2", 10],
        ["PROJ", "PROJ-6",  "EW", "api",  "epic-3",  "sum6",  "desc",  "Task", "u3", "2026-01-07", "P-003", "t6", "M-1003", "Dev", "acc2",  5],
        ["PROJ", "PROJ-7",  "J6", "dash", "epic-4",  "sum7",  "desc",  "Task", "u4", "2026-01-08", "P-004", "t7", "M-1004", "Dev", "acc2",  7],
        ["PROJ", "PROJ-8",  "JX", "sso",  "epic-5",  "sum8",  "desc",  "Task", "u1", "2026-01-09", "P-005", "t8", "M-1001", "Dev", "acc1",  2],
        ["PROJ", "PROJ-9",  "HW", "alrt", "epic-6",  "sum9",  "desc",  "Task", "u5", "2026-01-10", "P-006", "t9", "M-1005", "Dev", "acc3",  6],
        ["PROJ", "PROJ-10", "DF", "back", "epic-7",  "sum10", "desc",  "Task", "u1", "2026-01-11", "P-007", "t10","M-1001", "Dev", "acc1", 20],
        ["PROJ", "PROJ-11", "HE", "sec",  "epic-8",  "sum11", "desc",  "Task", "u2", "2026-01-12", "P-008", "t11","M-1002", "Dev", "acc1", 15],
        ["PROJ", "PROJ-12", "EW", "perf", "epic-9",  "sum12", "desc",  "Task", "u3", "2026-01-13", "P-009", "t12","M-1003", "Dev", "acc2",  8],
        ["PROJ", "PROJ-13", "PE", "tool", "epic-10", "sum13", "desc",  "Task", "u6", "2026-01-14", "P-010", "t13","M-1006", "Dev", "acc3",  4],
        ["PROJ", "PROJ-14", "WH", "test", "epic-12", "sum14", "desc",  "Task", "u2", "2026-01-15", "P-012", "t14","M-1002", "Dev", "acc1",  3],
        ["PROJ", "PROJ-15", "PF", "imp",  "epic-13", "sum15", "desc",  "Task", "u7", "2026-01-16", "P-013", "t15","M-1008", "Dev", "acc2",  1],
        ["PROJ", "PROJ-16", "KE", "exp",  "epic-14", "sum16", "desc",  "Task", "u6", "2026-01-17", "P-014", "t16","M-1006", "Dev", "acc3",  9],
    ]
    for r in rows:
        ws.append(r)

    # Filas de regresion v1.6.2 — Component Name y Matricula como STRING,
    # cruzan con peticiones/recursos NUMERIC del fichero del perfil Cierre:
    #   55751  + 99642 + Dev -> PROJ-20 (7) = 7
    #   101770 + 90014 + Dev -> PROJ-21 (3) + PROJ-22 (2) = 5
    #   138074 + 99641 + Dev -> PROJ-23 (9) = 9   (PROJ-24 con Sup excluido)
    regression_rows = [
        ["PROJ", "PROJ-20", "DF", "reg",  "epic-r1", "sumR1", "desc", "Task", "u1", "2026-02-01", "55751",  "tR1", "99642", "Dev", "accR", 7],
        ["PROJ", "PROJ-21", "EW", "reg",  "epic-r2", "sumR2", "desc", "Task", "u2", "2026-02-02", "101770", "tR2", "90014", "Dev", "accR", 3],
        ["PROJ", "PROJ-22", "EW", "reg",  "epic-r2", "sumR3", "desc", "Task", "u2", "2026-02-03", "101770", "tR3", "90014", "Dev", "accR", 2],
        ["PROJ", "PROJ-23", "HE", "reg",  "epic-r3", "sumR4", "desc", "Task", "u3", "2026-02-04", "138074", "tR4", "99641", "Dev", "accR", 9],
        ["PROJ", "PROJ-24", "HE", "reg",  "epic-r3", "sumR5", "desc", "Task", "u3", "2026-02-05", "138074", "tR5", "99641", "Sup", "accR", 4],
        # v1.8.1: imputacion que cruza con la fila P-016/M-1010 del fichero
        # del perfil Cierre (cuyo Usuario_Resp_Tecnico viene con padding
        # "MG002   "). Funcion=Dev, Hours=5. Tras el trim, el SUMIFS de
        # Jira casa -> 5h, y la celda ("M-1010", "MG002") de la tabla por
        # responsable de Resumen vale PDCL=6.0.
        ["PROJ", "PROJ-25", "DF", "pad",  "epic-pad", "sumR6", "desc", "Task", "u8", "2026-02-10", "P-016", "tR6", "M-1010", "Dev", "accR", 5],
    ]
    for r in regression_rows:
        ws.append(r)

    # Filas huerfanas v1.7.0 — imputaciones cuya (Component Name, Matricula)
    # NO casa con ninguna (Peticion, Recurso) del fichero del perfil Cierre.
    # Se usan para verificar que el builder las incluye en Resultado como
    # filas adicionales (opt-in mes.orphans.enabled=true).
    # Totales esperados por pareja (CN, Mat):
    #   (TICKETS, -)            -> 2 imputaciones x 4h = 8h
    #   (VACACIONES, 90014)     -> 3h   (90014 existe en el fichero Cierre pero asociado a 101770, no a VACACIONES)
    #   (P-001, MAT-HUERFANO)   -> 1h   (P-001 existe en el fichero Cierre pero solo con M-1001; MAT-HUERFANO no esta asociado)
    orphan_rows = [
        ["PROJ", "PROJ-30", "PF", "huerf", "epic-h1", "sumH1", "desc", "Task", "u4", "2026-02-10", "TICKETS",      "tH1", "-",             "Dev", "accH", 4],
        ["PROJ", "PROJ-31", "PF", "huerf", "epic-h1", "sumH2", "desc", "Task", "u4", "2026-02-11", "TICKETS",      "tH2", "-",             "Dev", "accH", 4],
        ["PROJ", "PROJ-32", "EW", "huerf", "epic-h2", "sumH3", "desc", "Task", "u5", "2026-02-12", "VACACIONES",   "tH3", "90014",         "Dev", "accH", 3],
        ["PROJ", "PROJ-33", "DF", "huerf", "epic-h3", "sumH4", "desc", "Task", "u1", "2026-02-13", "P-001",        "tH4", "MAT-HUERFANO",  "Dev", "accH", 1],
    ]
    for r in orphan_rows:
        ws.append(r)

    wb.save(f"{OUT_DIR}/extraccion.xlsx")
    print("Generated extraccion.xlsx (perfil Extraccion v2.0.0 = export Jira): "
          "1 meta + 1 header + 16 historical + 5 regression v1.6.2 + "
          "1 regression v1.8.1 + 4 orphan v1.7.0 = 28 filas")


# =========================================================================
# deuda.xlsx — horas de deuda opcionales (v2.2.0)
# =========================================================================
# Perfil Deuda (v2.2.0): fichero OPCIONAL que aporta horas de deuda por
# (Peticion, Matricula, Funcion). Cabeceras en fila 1. Cuando esta
# presente, la columna "PDCL + Deuda" de la hoja Resultado suma las
# horas cruzadas. Cuando no, la columna es identica a "PDCL".
#
# Escenarios cubiertos en el fixture (documentados para los tests):
#
#   Fila A: (P-001, M-1001, Dev, 5h)
#     -> cruza con la fila P-001 del fichero Cierre (Matricula=M-1001,
#        Res. Tecnico=Dev).
#     -> Fila P-001 en Resultado: PDCL = Jira*1.2, Jira = SUMIFS(...).
#        PDCL + Deuda = PDCL + 5.
#
#   Fila B: (P-002, M-1002, Dev, 3h)
#     -> cruza con P-002 del fichero Cierre.
#     -> PDCL + Deuda = PDCL(P-002) + 3.
#
#   Fila C: (P-007, M-1001, Dev, 10h)
#     -> cruza con P-007 (Matricula=M-1001, Res. Tecnico=Dev).
#     -> PDCL + Deuda = PDCL(P-007) + 10.
#
#   Fila D: (P-001, M-1001, Dev, 2h)
#     -> segunda entrada para la misma clave de la fila A.
#     -> Total deuda para (P-001, M-1001, Dev) = 5 + 2 = 7h.
#     -> PDCL + Deuda = PDCL(P-001) + 7.
#
#   Fila E: (P-999, M-9999, Dev, 100h)
#     -> NO casa con ninguna fila de Resultado (peticion inexistente).
#     -> El SUMIFS no la encuentra desde ninguna fila de Resultado; el
#        fixture sirve para verificar que las filas sin match en Deuda
#        devuelven SUMIFS=0 (no afectan al resto).
#
#   Fila F: (P-010, -, Dev, 4h)
#     -> fila con Matricula="-" (placeholder estilo v1.6.1). Aqui la
#        Peticion P-010 SI existe en Cierre pero con Matricula=M-1006,
#        no "-". Por tanto en Resultado la fila de P-010 tiene
#        Matricula=M-1006 y el SUMIFS(Deuda[Matricula="-"]) NO casa.
#        Sirve como canario: valor=0 para esa fila de Resultado,
#        confirma que el placeholder no produce un falso match.
#
# Totales esperados:
#   P-001 fila en Resultado -> deuda agregada = 5 + 2 = 7
#   P-002 fila en Resultado -> deuda agregada = 3
#   P-007 fila en Resultado -> deuda agregada = 10
#   resto de filas de Resultado -> deuda agregada = 0
DEUDA_PROFILE_HEADERS = ["Peticion", "Matricula", "Funcion", "Horas"]


def build_deuda_profile_fixture():
    """Genera deuda.xlsx (perfil Deuda v2.2.0 = horas de deuda opcional)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Deuda"
    ws.append(DEUDA_PROFILE_HEADERS)

    rows = [
        # (A) cruza con P-001 / M-1001 / Dev
        ["P-001", "M-1001", "Dev", 5],
        # (B) cruza con P-002 / M-1002 / Dev
        ["P-002", "M-1002", "Dev", 3],
        # (C) cruza con P-007 / M-1001 / Dev
        ["P-007", "M-1001", "Dev", 10],
        # (D) segunda entrada para la misma clave que (A) -> se agrega
        ["P-001", "M-1001", "Dev", 2],
        # (E) peticion inexistente en Cierre -> no afecta a Resultado
        ["P-999", "M-9999", "Dev", 100],
        # (F) Matricula="-" (placeholder v1.6.1); Peticion P-010 si existe
        #     en Cierre pero con Matricula=M-1006, no "-". No casa.
        ["P-010", "-", "Dev", 4],
    ]
    for r in rows:
        ws.append(r)

    wb.save(f"{OUT_DIR}/deuda.xlsx")
    print("Generated deuda.xlsx (perfil Deuda v2.2.0 = horas de deuda): "
          "1 header + 6 filas (incluye match multiple, no-match, y placeholder '-').")


if __name__ == "__main__":
    import os
    os.makedirs(OUT_DIR, exist_ok=True)
    build_cierre_profile_fixture()
    build_extraccion_profile_fixture()
    build_deuda_profile_fixture()
